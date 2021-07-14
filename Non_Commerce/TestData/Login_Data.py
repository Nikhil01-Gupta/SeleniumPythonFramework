import openpyxl

from configparser import ConfigParser

from Non_Commerce.TestSuites.Test_Scenarion_Data import Test_Scenario_Data


class TestLogin_Data:
    config = ConfigParser()
    config.read(r'D:\PycharmProjects\MyThird_Project\Non_Commerce\Config_example.cfg')
    TestData_File = config['TestData']['TestData_location']

    @staticmethod
    def getLogin_Data():
        TestScanerio_Obj = Test_Scenario_Data.getscenario()
        DataFile_Name = TestLogin_Data.TestData_File

        #workbook = openpyxl.load_workbook("D://ABFL_Python_TestData//TestData//"+DataFile_Name)
        workbook = openpyxl.load_workbook("D:\PycharmProjects\MyThird_Project\Test_Data\WebAutomationTestData.xlsx")
        sheet = workbook['DT_LoginEntity']
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=2).value == TestScanerio_Obj['TestCase']:
                for j in range(3, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]



TestLogin_Data.getLogin_Data()