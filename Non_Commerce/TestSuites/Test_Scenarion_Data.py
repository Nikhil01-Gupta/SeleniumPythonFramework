import openpyxl
from configparser import ConfigParser



class Test_Scenario_Data:
   config = ConfigParser()
   config.read(r'D:\PycharmProjects\MyThird_Project\Non_Commerce\Config_example.cfg')
   test_scenario_path=config['TestScanerio']['TestScanerio_location']

   @staticmethod
   def getscenario():
      TestScanerio = Test_Scenario_Data.test_scenario_path
      #print("properly working"+TestScanerio)
      #workbook = openpyxl.load_workbook("C:\\Users\\Me\\PycharmProjects\\MyFirst_Project\\TestSuites\\"+TestScanerio)
      get_workbook = openpyxl.load_workbook(r'D:\PycharmProjects\MyThird_Project\TestSuites\1000_TestScenarionTestSuites.xlsx')
      sheet = get_workbook['TestScenarion']
      ColNames = {}
      Current = 0
      Suite=[]
      All_Scanerios = []
      for col in sheet.iter_cols(1,sheet.max_column):
         ColNames[col[0].value]=Current
         Current += 1

      for row_cells in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
         if row_cells[ColNames['ExecuteTestScenario']].value == 'Yes':
            TestCaseName = row_cells[ColNames['Reference']].value
            TestCaseScenarioObjective = row_cells[ColNames['Scenario Objective']].value
            TestCaseDescription = row_cells[ColNames['Description']].value
            TestCaseScripter = row_cells[ColNames['AutomationScripterName']].value
            Suite.append('TestCase')
            Suite.append(TestCaseName)
            Suite.append('TestObjective')
            Suite.append(TestCaseScenarioObjective)
            Suite.append('TestDescription')
            Suite.append(TestCaseDescription)
            Suite.append('TestScripter')
            Suite.append(TestCaseScripter)

            def Convert(lst):
               res_dct = {lst[i]: lst[i + 1] for i in range(0, len(lst), 2)}
               return res_dct

            All_Scanerios.append(Convert(Suite))

      for x in range(0, len(All_Scanerios)):
         # print(All_Scanerios[x])
         return All_Scanerios[x]







Test_Scenario_Data.getscenario()




