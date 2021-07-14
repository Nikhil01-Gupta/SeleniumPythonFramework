import configparser
import os

import openpyxl
import pytest
from configparser import ConfigParser
from selenium import webdriver
from datetime import datetime
import time

from Non_Commerce.TestSuites.Test_Scenarion_Data import Test_Scenario_Data
from Non_Commerce.utilities.BaseClass import BaseClass

config = ConfigParser()
confi_File_Path = r"D:\PycharmProjects\MyThird_Project\Non_Commerce\Config.ini"
config.read(confi_File_Path)
HtmlReport = config["Report"]["Generate_HTML_Report"]
base = BaseClass()
log = base.get_logger()

config1 = configparser.ConfigParser()
config1.read(r"D:\PycharmProjects\MyThird_Project\Non_Commerce\Config.ini")


@pytest.fixture(scope="class")
def setup(request):
    global driver
    Current = 0
    ColNames = {}
    Excel_PATH = config1["general"]["default_path"]
    get_workbook = openpyxl.load_workbook(Excel_PATH)
    Sheet = get_workbook['Configuration']
    for col in Sheet.iter_cols(1, Sheet.max_column):
        ColNames[col[0].value] = Current
        Current += 1

    for row_cells in Sheet.iter_rows(min_row=1, max_row=Sheet.max_row):
        browser_name = row_cells[ColNames['Browser']].value

    #browser_name = config["BrowserType"]["BrowserName"]
    #config.get('BrowserType','BrowserName')
    if browser_name == "Chrome":
        driver = webdriver.Chrome(executable_path=r"D:\chrome_setup\chromedriver.exe")
    elif browser_name == "Firefox":
        driver = webdriver.Firefox(executable_path=r"D:\chrome_setup\geckodriver.exe")
    driver.get("https://admin-demo.nopcommerce.com")
    driver.maximize_window()
    time.sleep(3)
    request.cls.driver = driver


    #driver.maximize_window()
    #request.cls.driver = driver
    #return driver
    yield
    driver.close()



if HtmlReport == 'Yes':
    @pytest.hookimpl(tryfirst=True)
    def pytest_configure(config):
        if not os.path.exists('D://PycharmProjects//MyThird_Project//HTML_REPORT'):
            os.makedirs('D://PycharmProjects//MyThird_Project//HTML_REPORT')

        TestScanerio_Obj = str(Test_Scenario_Data.getscenario()['TestCase'])
        date_TestCase = datetime.now().strftime("%d-%m-%Y %I-%M-%S_%p")+" " + TestScanerio_Obj
        config.option.htmlpath = 'D://PycharmProjects//MyThird_Project//HTML_REPORT//'+date_TestCase+'//'+date_TestCase+".html"

@pytest.mark.hookwrapper
def pytest_runtest_makereport(item):
    """
        Extends the PyTest Plugin to take and embed screenshot in html report, whenever test fails.
        :param item:
        """

    TestScanerio_Obj = str(Test_Scenario_Data.getscenario()['TestCase'])
    date_TestCase = datetime.now().strftime("%d-%m-%Y %I-%M-%S_%p") + " " + TestScanerio_Obj
    directory = 'D://PycharmProjects//MyThird_Project//HTML_REPORT//' + date_TestCase + '//' + date_TestCase
    pytest_html = item.config.pluginmanager.getplugin('html')
    outcome = yield
    report = outcome.get_result()
    extra = getattr(report, 'extra', [])

    if report.when == 'call' or report.when == "setup":
        xfail = hasattr(report, 'wasxfail')
        if (report.skipped and xfail) or (report.failed and not xfail):
            file_name = report.nodeid.replace("::", "_") + ".png"
            _capture_screenshot(file_name)
            if file_name:
                html = '<div><img src="%s" alt="screenshot" style="width:304px;height:228px;" ' \
                       'onclick="window.open(this.src)" align="right"/></div>' % file_name
                extra.append(pytest_html.extras.html(html))
        report.extra = extra


def _capture_screenshot(name):
    TestScanerio_Obj = str(Test_Scenario_Data.getscenario()['TestCase'])
    date_TestCase = datetime.now().strftime("%d-%m-%Y %I-%M-%S_%p") + " " + TestScanerio_Obj
    directory = 'D://PycharmProjects//MyThird_Project//HTML_REPORT//' + date_TestCase + '//' + date_TestCase
    driver.get_screenshot_as_file(name)